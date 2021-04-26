from openpyxl import Workbook
from openpyxl import load_workbook

def write_to_excel(data):
    ''' function to write data to a new excel file '''
    wb = Workbook()
    ws = wb.active

    # write data to the rows as long as there is data
    for row, url in zip(ws.iter_rows(max_row=len(data)), data):
        for cell in row:
            cell.value = url

    wb.save('urls.xlsx')
    print('--- done writing to excel ---')

def append_to_excel(data):
    ''' function to append data to an existing excel file '''
    wb = load_workbook('urls.xlsx')
    ws = wb.active

    # ws.max_row returns the last row with data written in it
    start = ws.max_row
    # write data to rows that start after the last already existing row as long as there is data
    for row, url in zip(ws.iter_rows(min_row=start, max_row=start+len(data), max_col=1), data):
        for cell in row:
            cell.value = url

    wb.save('urls.xlsx')
    print('--- done appending to excel ---')